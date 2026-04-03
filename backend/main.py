import os
import io
import uuid
import json
import logging
from typing import List, Optional, Dict, Any
from datetime import datetime
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Request, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
from langchain_openai import AzureChatOpenAI, AzureOpenAIEmbeddings, ChatOpenAI, OpenAIEmbeddings
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage
from tavily import TavilyClient
from pymilvus import MilvusClient, Collection, connections, utility, FieldSchema, CollectionSchema, DataType
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import re
from rank_bm25 import BM25Okapi
import jieba

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

load_dotenv(override=True)

app = FastAPI(title="Knowledge Base API", version="1.0.0")

@app.on_event("startup")
async def startup_event():
    """应用启动时加载 BM25 索引"""
    logger.info("正在加载 BM25 索引...")
    bm25_index.load()
    logger.info(f"BM25 索引加载完成，文档数: {len(bm25_index.documents)}")

llm = AzureChatOpenAI(
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("AZURE_OPENAI_API_VERSION"),
    deployment_name=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"),
    temperature=0.7,
)

embeddings = AzureOpenAIEmbeddings(
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("AZURE_OPENAI_API_VERSION"),
    model=os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT_NAME"),
)

EMBEDDING_DIM_MAP = {
    "text-embedding-3-small": 1536,
    "text-embedding-3-large": 3072,
    "text-embedding-ada-002": 1536,
    "BAAI/bge-large-zh-v1.5": 1024,
    "BAAI/bge-base-zh-v1.5": 768,
    "BAAI/bge-small-zh-v1.5": 512,
    "BAAI/bge-m3": 1024,
    "sentence-transformers/all-MiniLM-L6-v2": 384,
    "text-embedding-v3": 1024,
    "text-embedding-v2": 1536,
    "text-embedding-async-v2": 1536,
}

def create_embeddings(config: dict):
    """根据前端传入的 embedding_config 动态创建 embeddings 实例"""
    if not config:
        return embeddings
    provider = config.get("provider", "AzureOpenAI")
    if provider == "AzureOpenAI":
        return AzureOpenAIEmbeddings(
            azure_endpoint=config["endpoint"],
            api_key=config["api_key"],
            api_version=config["api_version"],
            model=config.get("deployment", config.get("model")),
        )
    else:
        return OpenAIEmbeddings(
            api_key=config["api_key"],
            base_url=config.get("base_url"),
            model=config["model"],
        )

def get_embedding_dim(config: dict) -> int:
    """根据 embedding_config 推断嵌入维度"""
    model = config.get("model", "")
    # Azure 可能有 deployment 名称映射到实际 model
    deployment = config.get("deployment", "")
    dim = EMBEDDING_DIM_MAP.get(model) or EMBEDDING_DIM_MAP.get(deployment)
    if dim:
        return dim
    # 默认回退到 1536
    logger.warning(f"未知嵌入模型 '{model}'，使用默认维度 1536")
    return 1536

tavily_client = TavilyClient(api_key=os.getenv("TAVILY_API_KEY"))

MILVUS_HOST = os.getenv("MILVUS_HOST", "localhost")
MILVUS_PORT = os.getenv("MILVUS_PORT", "19530")

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

DOCUMENTS_FILE = os.path.join(os.path.dirname(__file__), "documents.json")

class Message(BaseModel):
    role: str
    content: str

class ModelConfig(BaseModel):
    """动态模型配置"""
    provider: str = "AzureOpenAI"
    model: str = ""
    api_key: str = ""
    base_url: Optional[str] = None
    endpoint: Optional[str] = None
    api_version: Optional[str] = None
    temperature: float = 0.7

class ChatRequest(BaseModel):
    messages: List[Message]
    system_prompt: str = "You are a helpful assistant."
    top_k: int = 5
    use_hybrid: bool = True
    use_query_expansion: bool = True
    web_search_enabled: bool = True
    llm_config: Optional[Dict[str, Any]] = None  # 动态模型配置
    embedding_config: Optional[Dict[str, Any]] = None  # 动态嵌入配置

class ChatResponse(BaseModel):
    response: str
    search_results: Optional[List[dict]] = None
    knowledge_sources: Optional[List[dict]] = None
    used_knowledge: bool = False

class DocumentInfo(BaseModel):
    id: str
    filename: str
    upload_time: str
    file_size: int
    file_type: str
    status: str
    chunks_count: int = 0
    error: Optional[str] = None  # 新增错误信息字段

BM25_INDEX_FILE = os.path.join(os.path.dirname(__file__), "bm25_index.json")

class BM25Index:
    def __init__(self):
        self.bm25 = None
        self.documents = []
        self.tokenized_corpus = []
    
    def add_documents(self, documents: List[Dict[str, Any]]):
        for doc in documents:
            self.documents.append(doc)
            tokenized = list(jieba.cut(doc['content']))
            self.tokenized_corpus.append(tokenized)
        if self.tokenized_corpus:
            self.bm25 = BM25Okapi(self.tokenized_corpus)
    
    def search(self, query: str, top_k: int = 5) -> List[Dict[str, Any]]:
        if self.bm25 is None:
            return []
        tokenized_query = list(jieba.cut(query))
        scores = self.bm25.get_scores(tokenized_query)
        top_indices = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)[:top_k]
        results = []
        for idx in top_indices:
            if scores[idx] > 0:
                result = self.documents[idx].copy()
                result['score'] = float(scores[idx])
                results.append(result)
        return results
    
    def clear(self):
        self.bm25 = None
        self.documents = []
        self.tokenized_corpus = []
    
    def save(self):
        data = {
            'documents': self.documents,
            'tokenized_corpus': self.tokenized_corpus
        }
        with open(BM25_INDEX_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    
    def load(self):
        if os.path.exists(BM25_INDEX_FILE):
            with open(BM25_INDEX_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.documents = data.get('documents', [])
            self.tokenized_corpus = data.get('tokenized_corpus', [])
            if self.tokenized_corpus:
                self.bm25 = BM25Okapi(self.tokenized_corpus)

bm25_index = BM25Index()

_milvus_collections: Dict[int, Collection] = {}
_milvus_connected = False

def _connect_milvus():
    global _milvus_connected
    if _milvus_connected:
        return
    logger.info(f"尝试连接Milvus: {MILVUS_HOST}:{MILVUS_PORT}")
    try:
        connections.connect(host=MILVUS_HOST, port=MILVUS_PORT)
        _milvus_connected = True
        logger.info("Milvus连接成功")
    except Exception as e:
        logger.error(f"Milvus连接失败: {str(e)}", exc_info=True)
        raise

def get_or_create_collection(dim: int = 1536) -> Collection:
    """按维度获取或创建 Milvus collection，使用全局缓存"""
    global _milvus_collections

    if dim in _milvus_collections:
        return _milvus_collections[dim]

    _connect_milvus()

    collection_name = f"knowledge_base_{dim}"

    if utility.has_collection(collection_name):
        collection = Collection(collection_name)
        logger.info(f"使用已存在的集合: {collection_name}")
    else:
        logger.info(f"创建新集合: {collection_name} (dim={dim})")
        fields = [
            FieldSchema(name="id", dtype=DataType.VARCHAR, is_primary=True, max_length=100),
            FieldSchema(name="document_id", dtype=DataType.VARCHAR, max_length=100),
            FieldSchema(name="content", dtype=DataType.VARCHAR, max_length=65535),
            FieldSchema(name="embedding", dtype=DataType.FLOAT_VECTOR, dim=dim),
            FieldSchema(name="metadata", dtype=DataType.VARCHAR, max_length=65535),
        ]
        schema = CollectionSchema(fields, description="Knowledge base collection")
        collection = Collection(name=collection_name, schema=schema)
        collection.create_index(field_name="embedding", index_params={"index_type": "IVF_FLAT", "metric_type": "COSINE", "params": {"nlist": 128}})
        logger.info(f"集合 {collection_name} 创建成功并添加索引")

    _milvus_collections[dim] = collection
    return collection

# 向后兼容：默认使用 1536 维的 collection
def get_milvus_collection():
    return get_or_create_collection(1536)

def load_documents():
    if os.path.exists(DOCUMENTS_FILE):
        with open(DOCUMENTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_documents(documents):
    with open(DOCUMENTS_FILE, 'w', encoding='utf-8') as f:
        json.dump(documents, f, ensure_ascii=False, indent=2)

def extract_text_from_pdf(file_content: bytes) -> str:
    pdf_file = io.BytesIO(file_content)
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for i, page in enumerate(pdf_reader.pages):
        page_text = page.extract_text()
        text += page_text + "\n"
        logger.info(f"PDF第{i+1}页提取文本长度: {len(page_text)}")
    
    logger.info(f"PDF总文本长度: {len(text)}, 前100字符: {text[:100]}")
    
    if len(text.strip()) < 50:
        logger.warning("PDF提取的文本过少，可能是图片型PDF，建议使用OCR工具或转换格式")
    
    return text

def extract_text_from_docx(file_content: bytes) -> str:
    docx_file = io.BytesIO(file_content)
    doc = Document(docx_file)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text

def extract_text_from_txt(file_content: bytes) -> str:
    return file_content.decode('utf-8', errors='ignore')

def extract_text_from_excel(file_content: bytes) -> str:
    excel_file = io.BytesIO(file_content)
    workbook = load_workbook(excel_file, data_only=True)
    text = ""
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text += f"工作表: {sheet_name}\n"
        
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text += row_text + "\n"
    
    logger.info(f"Excel文本提取完成，长度: {len(text)}")
    return text

def split_text_into_chunks(text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
    sentences = re.split(r'(?<=[。！？\n])', text)
    chunks = []
    current_chunk = ""
    
    for sentence in sentences:
        if len(current_chunk) + len(sentence) < chunk_size:
            current_chunk += sentence
        else:
            if current_chunk:
                chunks.append(current_chunk.strip())
            current_chunk = sentence
    
    if current_chunk:
        chunks.append(current_chunk.strip())
    
    return [chunk for chunk in chunks if len(chunk) > 50]

def process_document(document_id: str, filename: str, file_content: bytes, file_type: str,
                     chunk_size: int = 500, overlap: int = 50, doc_embeddings=None, dim: int = 1536):
    logger.info(f"开始处理文档: {filename} (ID: {document_id}, 类型: {file_type}, 维度: {dim})")
    try:
        logger.info(f"步骤1: 提取文本内容...")
        if file_type == "application/pdf":
            text = extract_text_from_pdf(file_content)
            logger.info(f"PDF文本提取完成，长度: {len(text)}")
        elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            text = extract_text_from_docx(file_content)
            logger.info(f"DOCX文本提取完成，长度: {len(text)}")
        elif file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            text = extract_text_from_excel(file_content)
            logger.info(f"Excel文本提取完成，长度: {len(text)}")
        elif file_type == "text/plain":
            text = extract_text_from_txt(file_content)
            logger.info(f"TXT文本提取完成，长度: {len(text)}")
        else:
            error_msg = f"不支持的文件类型: {file_type}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        if not text or len(text.strip()) == 0:
            error_msg = "提取的文本内容为空"
            logger.error(error_msg)
            return False, error_msg

        logger.info(f"步骤2: 分割文本为块 (chunk_size={chunk_size}, overlap={overlap})...")
        chunks = split_text_into_chunks(text, chunk_size, overlap)
        logger.info(f"文本分割完成，生成 {len(chunks)} 个文本块")

        if len(chunks) == 0:
            error_msg = "未能生成有效的文本块"
            logger.error(error_msg)
            return False, error_msg

        logger.info(f"步骤3: 获取Milvus集合 (dim={dim})...")
        collection = get_or_create_collection(dim)
        logger.info(f"Milvus集合获取成功")

        logger.info(f"步骤4: 批量生成嵌入向量...")

        # 使用传入的 embeddings 实例或全局默认
        emb = doc_embeddings or embeddings
        all_embeddings = emb.embed_documents(chunks)
        logger.info(f"嵌入向量批量生成完成，共 {len(all_embeddings)} 个")
        
        data = []
        for i, (chunk, embedding) in enumerate(zip(chunks, all_embeddings)):
            chunk_id = f"{document_id}_{i}"
            metadata = json.dumps({
                "filename": filename,
                "chunk_index": i,
                "total_chunks": len(chunks)
            }, ensure_ascii=False)
            
            data.append({
                "id": chunk_id,
                "document_id": document_id,
                "content": chunk,
                "embedding": embedding,
                "metadata": metadata
            })
        
        logger.info(f"所有嵌入向量生成完成，开始插入Milvus...")
        collection.insert(data)
        collection.flush()
        logger.info(f"数据插入Milvus成功")
        
        logger.info(f"步骤5: 更新BM25索引...")
        bm25_docs = []
        for i, chunk in enumerate(chunks):
            chunk_id = f"{document_id}_{i}"
            metadata = {
                "filename": filename,
                "chunk_index": i,
                "total_chunks": len(chunks)
            }
            bm25_docs.append({
                "id": chunk_id,
                "document_id": document_id,
                "content": chunk,
                "metadata": metadata
            })
        bm25_index.add_documents(bm25_docs)
        bm25_index.save()
        logger.info(f"BM25索引更新成功")
        
        logger.info(f"文档处理成功: {filename} (ID: {document_id})")
        return True, len(chunks)
    except Exception as e:
        error_msg = f"文档处理失败: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return False, error_msg

def expand_query(query: str, llm_instance=None) -> List[str]:
    check_llm = llm_instance or llm
    expansion_prompt = f"""为以下查询生成3个语义相近的变体，用于提高检索召回率。
    原查询：{query}

    请以JSON格式返回，只包含数组，不要其他内容：
    ["变体1", "变体2", "变体3"]"""
    try:
        response = check_llm.invoke([HumanMessage(content=expansion_prompt)])
        content = response.content.strip()
        if content.startswith('[') and content.endswith(']'):
            queries = json.loads(content)
            if isinstance(queries, list) and len(queries) > 0:
                logger.info(f"查询扩展成功: {queries}")
                return queries
        return [query]
    except Exception as e:
        logger.warning(f"查询扩展失败: {str(e)}，使用原查询")
        return [query]

def reciprocal_rank_fusion(vector_results: List[Dict], bm25_results: List[Dict], k: int = 60) -> List[Dict]:
    rrf_scores = {}
    
    for rank, result in enumerate(vector_results):
        doc_id = result['id']
        score = 1.0 / (k + rank + 1)
        if doc_id not in rrf_scores:
            rrf_scores[doc_id] = {'score': 0, 'result': result}
        rrf_scores[doc_id]['score'] += score
    
    for rank, result in enumerate(bm25_results):
        doc_id = result['id']
        score = 1.0 / (k + rank + 1)
        if doc_id not in rrf_scores:
            rrf_scores[doc_id] = {'score': 0, 'result': result}
        rrf_scores[doc_id]['score'] += score
    
    sorted_results = sorted(rrf_scores.values(), key=lambda x: x['score'], reverse=True)
    return [item['result'] for item in sorted_results]

def search_knowledge_base(query: str, top_k: int = 5, use_hybrid: bool = True, use_query_expansion: bool = True,
                           search_embeddings=None, dim: int = 1536, llm_instance=None) -> List[Dict[str, Any]]:
    try:
        collection = get_or_create_collection(dim)
        collection.load()

        queries = [query]
        if use_query_expansion:
            queries = expand_query(query, llm_instance=llm_instance)

        logger.info(f"开始混合检索，查询数量: {len(queries)}, top_k: {top_k}, dim={dim}")

        emb = search_embeddings or embeddings
        vector_results = []
        for q in queries:
            query_embedding = emb.embed_query(q)
            logger.info(f"查询: {q}, 嵌入向量维度: {len(query_embedding)}")
            
            search_params = {"metric_type": "COSINE", "params": {"nprobe": 16}}
            results = collection.search(
                data=[query_embedding],
                anns_field="embedding",
                param=search_params,
                limit=top_k * 2,
                output_fields=["content", "document_id", "metadata"]
            )
            
            for result in results[0]:
                try:
                    metadata_str = result.get("metadata")
                    metadata = json.loads(metadata_str) if metadata_str else {}
                except:
                    metadata = {}
                
                vector_results.append({
                    "id": result.id,
                    "content": result.get("content"),
                    "score": float(result.score),
                    "filename": metadata.get("filename", ""),
                    "chunk_index": metadata.get("chunk_index", 0)
                })
        
        bm25_results = []
        if use_hybrid:
            for q in queries:
                bm25_hits = bm25_index.search(q, top_k * 2)
                for hit in bm25_hits:
                    bm25_results.append({
                        "id": hit['id'],
                        "content": hit['content'],
                        "score": hit['score'],
                        "filename": hit['metadata'].get('filename', ''),
                        "chunk_index": hit['metadata'].get('chunk_index', 0)
                    })
        
        if use_hybrid and bm25_results:
            knowledge_sources = reciprocal_rank_fusion(vector_results, bm25_results)[:top_k]
        else:
            deduplicated = {}
            for result in vector_results:
                if result['id'] not in deduplicated:
                    deduplicated[result['id']] = result
            knowledge_sources = list(deduplicated.values())[:top_k]
        
        logger.info(f"知识库搜索结果: 找到 {len(knowledge_sources)} 个匹配")
        for i, source in enumerate(knowledge_sources):
            logger.info(f"  结果{i+1}: 分数={source['score']:.4f}, 文件={source['filename']}, 内容={source['content'][:50]}...")
        
        return knowledge_sources
    except Exception as e:
        logger.error(f"知识库搜索失败: {str(e)}", exc_info=True)
        return []

def classify_query(query: str, llm_instance=None) -> str:
    """查询分类器：返回 kb_search / web_search / no_search"""
    check_llm = llm_instance or llm
    prompt = f"""将用户查询分类到以下类别之一：
- kb_search：需要从知识库中检索特定信息（如公司政策、产品文档、内部规定、已上传的资料）
- web_search：需要联网获取最新或实时信息（如新闻、天气、发布时间、价格、产品动态）
- no_search：常识性或创造性问题，无需额外检索即可回答（如编程、翻译、数学计算）

用户查询：{query}
仅回答类别名（kb_search / web_search / no_search），不要其他内容。"""
    try:
        response = check_llm.invoke([HumanMessage(content=prompt)])
        for label in ["kb_search", "web_search", "no_search"]:
            if label in response.content.lower():
                return label
    except Exception as e:
        logger.warning(f"查询分类失败: {e}")
    return "web_search"  # 默认走搜索，宁可多搜不漏

def is_knowledge_relevant(query: str, knowledge_sources: List[Dict[str, Any]], llm_instance=None) -> bool:
    if not knowledge_sources:
        return False

    top_score = knowledge_sources[0]["score"]

    logger.info(f"知识库最高相似度分数: {top_score:.4f}")

    # 分数过低，直接跳过
    if top_score < 0.3:
        return False

    # 用 LLM 判断检索到的内容是否真正包含与问题相关的信息
    check_llm = llm_instance or llm
    top_contents = [s["content"][:300] for s in knowledge_sources[:3] if s.get("content")]
    if not top_contents:
        return False

    check_prompt = f"""请判断以下知识库检索结果是否包含能够回答用户问题的信息。

用户问题：{query}

知识库检索结果：
{chr(10).join(f'{i+1}. {c}' for i, c in enumerate(top_contents))}

如果检索结果中包含与问题直接相关的信息或能帮助回答问题，回答"是"。
如果检索结果与问题无关，或只是一些通用/测试内容无法回答该问题，回答"否"。
只回答"是"或"否"，不要其他内容。"""

    try:
        response = check_llm.invoke([HumanMessage(content=check_prompt)])
        is_relevant = "是" in response.content and "否" not in response.content
        logger.info(f"LLM 知识相关性判断: {is_relevant} (问题: {query[:30]}...)")
        return is_relevant
    except Exception as e:
        logger.warning(f"LLM 相关性判断失败: {e}，回退到分数阈值判断")
        return top_score > 0.85

def perform_search(query: str) -> str:
    try:
        result = tavily_client.search(
            query=query,
            search_depth="basic",
            max_results=5,
            include_answer=True,
            include_raw_content=False
        )
        
        search_context = f"\n\n联网搜索结果：\n"
        if result.get("answer"):
            search_context += f"摘要：{result['answer']}\n\n"
        
        search_context += "参考来源：\n"
        for item in result.get("results", []):
            search_context += f"- {item.get('title', '')}\n  {item.get('url', '')}\n  {item.get('content', '')[:200]}...\n\n"
        
        return search_context, result.get("results", [])
    except Exception as e:
        return "", []

def create_llm(config: Optional[Dict[str, Any]] = None):
    """模型工厂函数：根据配置动态创建 LLM 实例"""
    if config is None:
        # 使用默认的 Azure OpenAI
        return llm
    
    provider = config.get("provider", "AzureOpenAI")
    
    if provider == "AzureOpenAI":
        return AzureChatOpenAI(
            azure_endpoint=config.get("endpoint") or os.getenv("AZURE_OPENAI_ENDPOINT"),
            api_key=config.get("api_key") or os.getenv("AZURE_OPENAI_API_KEY"),
            api_version=config.get("api_version") or os.getenv("AZURE_OPENAI_API_VERSION"),
            deployment_name=config.get("model") or os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"),
            temperature=config.get("temperature", 0.7),
        )
    else:
        # OpenAI 兼容接口（包括硅基流动、阿里百炼、DeepSeek 等）
        return ChatOpenAI(
            api_key=config.get("api_key"),
            base_url=config.get("base_url"),
            model=config.get("model"),
            temperature=config.get("temperature", 0.7),
        )

def prepare_chat_context(request: ChatRequest) -> tuple:
    """
    准备聊天上下文（消息构建、知识库搜索、上下文增强）
    Adaptive RAG 路由：先分类查询类型，再选择最优路径
    Returns: (messages, search_results, knowledge_sources, used_knowledge, current_llm)
    """
    # 在内部创建 LLM 实例，供所有辅助判断使用
    current_llm = create_llm(request.llm_config)

    messages = [SystemMessage(content=request.system_prompt)]

    for msg in request.messages:
        if msg.role == "user":
            messages.append(HumanMessage(content=msg.content))
        elif msg.role == "assistant":
            messages.append(AIMessage(content=msg.content))

    # 获取最后用户消息
    last_user_message = next(
        (msg.content for msg in reversed(request.messages) if msg.role == "user"),
        None
    )

    search_results = None
    knowledge_sources = None
    used_knowledge = False

    if last_user_message:
        logger.info(f"最后用户消息: {last_user_message}")

        # 确定搜索用的 embeddings 和维度
        search_emb = None
        search_dim = 1536
        if request.embedding_config:
            try:
                search_emb = create_embeddings(request.embedding_config)
                search_dim = get_embedding_dim(request.embedding_config)
            except Exception as e:
                logger.warning(f"创建搜索嵌入失败: {e}，使用默认配置")

        # Step 1: 查询分类
        route = classify_query(last_user_message, llm_instance=current_llm)
        logger.info(f"查询分类结果: {route}")

        allow_web_search = request.web_search_enabled

        if route == "kb_search":
            # 知识库检索路径
            knowledge_sources = search_knowledge_base(
                last_user_message,
                top_k=request.top_k,
                use_hybrid=request.use_hybrid,
                use_query_expansion=request.use_query_expansion,
                search_embeddings=search_emb,
                dim=search_dim,
                llm_instance=current_llm
            )

            if knowledge_sources and knowledge_sources[0]["score"] > 0.3 and is_knowledge_relevant(last_user_message, knowledge_sources, llm_instance=current_llm):
                used_knowledge = True
                logger.info(f"使用知识库内容，最高分数: {knowledge_sources[0]['score']:.4f}")
                knowledge_context = "\n\n知识库参考内容：\n"
                for i, source in enumerate(knowledge_sources[:min(3, request.top_k)], 1):
                    knowledge_context += f"{i}. {source['content']}\n"
                messages[-1] = HumanMessage(content=last_user_message + knowledge_context)
            else:
                # 知识库检索失败，降级到联网搜索
                logger.info("知识库检索未命中相关内容，降级到联网搜索")
                if not allow_web_search:
                    search_context, search_results = "", None
                else:
                    search_context, search_results = perform_search(last_user_message)
                if search_context:
                    messages[-1] = HumanMessage(content=last_user_message + search_context)
                else:
                    logger.info("联网搜索也无结果，直接回答")

        elif route == "web_search":
            # 联网搜索路径
            logger.info("执行联网搜索")
            if not allow_web_search:
                search_context, search_results = "", None
            else:
                search_context, search_results = perform_search(last_user_message)
            if search_context:
                messages[-1] = HumanMessage(content=last_user_message + search_context)
            else:
                logger.info("联网搜索无结果，直接回答")

        else:
            # no_search：直接回答，不增强上下文
            logger.info("常识性问题，直接回答")

    return messages, search_results, knowledge_sources, used_knowledge, current_llm

def process_document_background(document_id: str, filename: str, file_content: bytes,
                                 file_type: str, chunk_size: int, overlap: int,
                                 embedding_config: dict = None, dim: int = 1536):
    """后台异步处理文档"""
    try:
        logger.info(f"后台开始处理文档: {filename} (ID: {document_id}, dim={dim})")
        doc_embeddings = create_embeddings(embedding_config) if embedding_config else None
        success, result = process_document(document_id, filename, file_content, file_type, chunk_size, overlap, doc_embeddings, dim)
        
        documents = load_documents()
        for doc in documents:
            if doc["id"] == document_id:
                doc["status"] = "completed" if success else "failed"
                doc["chunks_count"] = result if success else 0
                doc["error"] = None if success else str(result)
                break
        save_documents(documents)
        logger.info(f"后台文档处理完成: {filename}, 状态: {'成功' if success else '失败'}")
    except Exception as e:
        logger.error(f"后台文档处理失败: {filename}, 错误: {e}")
        documents = load_documents()
        for doc in documents:
            if doc["id"] == document_id:
                doc["status"] = "failed"
                doc["error"] = str(e)
                break
        save_documents(documents)

@app.get("/")
async def root():
    return {"message": "Knowledge Base API is running"}

@app.get("/health")
async def health_check():
    try:
        collection = get_milvus_collection()
        milvus_status = "connected" if collection else "disconnected"
        return {"status": "healthy", "milvus": milvus_status}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

@app.get("/documents", response_model=List[DocumentInfo])
async def get_documents():
    documents = load_documents()
    return documents

@app.post("/upload")
async def upload_document(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    chunk_size: int = Form(500),
    overlap: int = Form(50),
    embedding_config: str = Form(None)
):
    """
    上传文档（异步处理）
    文件保存后立即返回，后台异步处理文档
    """
    logger.info(f"收到文件上传请求: {file.filename}, chunk_size={chunk_size}, overlap={overlap}")
    try:
        file_content = await file.read()
        file_size = len(file_content)
        logger.info(f"文件读取完成，大小: {file_size} bytes")

        # 解析嵌入配置
        emb_config = None
        emb_dim = 1536
        if embedding_config:
            try:
                emb_config = json.loads(embedding_config)
                emb_dim = get_embedding_dim(emb_config)
                logger.info(f"使用动态嵌入配置: model={emb_config.get('model')}, dim={emb_dim}")
            except json.JSONDecodeError:
                logger.warning("embedding_config JSON 解析失败，使用默认嵌入配置")

        document_id = str(uuid.uuid4())
        upload_time = datetime.now().isoformat()

        file_path = os.path.join(UPLOAD_DIR, f"{document_id}_{file.filename}")
        logger.info(f"保存文件到: {file_path}")
        with open(file_path, 'wb') as f:
            f.write(file_content)
        logger.info(f"文件保存成功")

        # 保存文档记录（状态为 processing）
        documents = load_documents()
        doc_record = {
            "id": document_id,
            "filename": file.filename,
            "upload_time": upload_time,
            "file_size": file_size,
            "file_type": file.content_type,
            "status": "processing",
            "chunks_count": 0,
            "error": None,
            "embedding_dimension": emb_dim,
        }
        if emb_config:
            # 保存 embedding_config，但移除 api_key 不落盘
            safe_config = {k: v for k, v in emb_config.items() if k != "api_key"}
            doc_record["embedding_config"] = safe_config
        documents.append(doc_record)
        save_documents(documents)
        logger.info(f"文档记录已添加，开始后台处理...")

        # 添加后台任务处理文档
        background_tasks.add_task(
            process_document_background,
            document_id, file.filename, file_content, file.content_type, chunk_size, overlap,
            emb_config, emb_dim
        )
        
        # 立即返回响应
        return {
            "success": True,
            "document_id": document_id,
            "filename": file.filename,
            "status": "processing",
            "message": "文档已上传，正在后台处理中"
        }
    except Exception as e:
        logger.error(f"文件上传异常: {str(e)}", exc_info=True)
        return {"success": False, "error": str(e)}

@app.get("/documents/{document_id}/status")
async def get_document_status(document_id: str):
    """查询文档处理状态"""
    documents = load_documents()
    for doc in documents:
        if doc["id"] == document_id:
            return {
                "id": doc["id"],
                "filename": doc["filename"],
                "status": doc["status"],
                "chunks_count": doc.get("chunks_count", 0),
                "error": doc.get("error")
            }
    raise HTTPException(status_code=404, detail="Document not found")

@app.delete("/documents/{document_id}")
async def delete_document(document_id: str):
    logger.info(f"收到删除文档请求: {document_id}")
    try:
        # 查找文档所在的维度
        documents = load_documents()
        target_doc = None
        for doc in documents:
            if doc["id"] == document_id:
                target_doc = doc
                break

        doc_dim = target_doc.get("embedding_dimension", 1536) if target_doc else 1536

        logger.info(f"步骤1: 获取Milvus集合 (dim={doc_dim})...")
        collection = get_or_create_collection(doc_dim)

        logger.info(f"步骤2: 加载Milvus集合...")
        collection.load()
        logger.info(f"Milvus集合加载成功")
        
        logger.info(f"步骤3: 从Milvus删除文档数据...")
        expr = f'document_id == "{document_id}"'
        collection.delete(expr)
        collection.flush()
        logger.info(f"Milvus数据删除成功")
        
        logger.info(f"步骤4: 从BM25索引删除文档数据...")
        # 同时过滤 documents 和 tokenized_corpus，保持索引对齐
        filtered_pairs = [
            (doc, tokens) 
            for doc, tokens in zip(bm25_index.documents, bm25_index.tokenized_corpus) 
            if doc['document_id'] != document_id
        ]
        if filtered_pairs:
            bm25_index.documents, bm25_index.tokenized_corpus = zip(*filtered_pairs)
            bm25_index.documents = list(bm25_index.documents)
            bm25_index.tokenized_corpus = list(bm25_index.tokenized_corpus)
        else:
            bm25_index.documents = []
            bm25_index.tokenized_corpus = []
        
        if bm25_index.tokenized_corpus:
            bm25_index.bm25 = BM25Okapi(bm25_index.tokenized_corpus)
        else:
            bm25_index.bm25 = None
        bm25_index.save()
        logger.info(f"BM25索引删除成功")
        
        logger.info(f"步骤5: 从documents.json删除文档记录...")
        documents = [doc for doc in documents if doc["id"] != document_id]
        save_documents(documents)
        logger.info(f"文档记录删除成功: {target_doc['filename'] if target_doc else 'unknown'}")
        
        logger.info(f"步骤5: 删除上传的文件...")
        deleted_files = []
        for filename in os.listdir(UPLOAD_DIR):
            if filename.startswith(document_id):
                file_path = os.path.join(UPLOAD_DIR, filename)
                os.remove(file_path)
                deleted_files.append(filename)
        logger.info(f"文件删除成功: {deleted_files}")
        
        logger.info(f"文档删除完成: {document_id}")
        return {"success": True, "message": "文档删除成功"}
    except Exception as e:
        logger.error(f"文档删除失败: {str(e)}", exc_info=True)
        return {"success": False, "error": str(e)}

@app.post("/chat", response_model=ChatResponse)
async def chat(request: ChatRequest):
    """聊天接口（非流式）"""
    try:
        logger.info(f"收到聊天请求，消息数: {len(request.messages)}")

        # 使用公共函数准备上下文（内部已创建 LLM 实例）
        messages, search_results, knowledge_sources, used_knowledge, current_llm = prepare_chat_context(request)

        response = current_llm.invoke(messages)
        logger.info(f"AI响应生成完成，使用知识库: {used_knowledge}")
        
        return ChatResponse(
            response=response.content,
            search_results=search_results,
            knowledge_sources=knowledge_sources[:3] if knowledge_sources else None,
            used_knowledge=used_knowledge
        )
    except Exception as e:
        logger.error(f"聊天处理失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/chat/stream")
async def chat_stream(request: ChatRequest, http_request: Request):
    """聊天接口（流式）"""
    async def generate():
        try:
            logger.info(f"收到流式聊天请求，消息数: {len(request.messages)}")
            
            # 使用公共函数准备上下文（内部已创建 LLM 实例）
            messages, search_results, knowledge_sources, used_knowledge, current_llm = prepare_chat_context(request)
            
            metadata = {
                "search_results": search_results,
                "knowledge_sources": knowledge_sources[:3] if knowledge_sources else None,
                "used_knowledge": used_knowledge
            }
            yield f"data: {json.dumps({'type': 'metadata', 'data': metadata}, ensure_ascii=False)}\n\n"
            
            full_response = ""
            async for chunk in current_llm.astream(messages):
                if chunk.content:
                    full_response += chunk.content
                    yield f"data: {json.dumps({'type': 'content', 'content': chunk.content}, ensure_ascii=False)}\n\n"
            
            logger.info(f"流式AI响应生成完成，使用知识库: {used_knowledge}")
            yield f"data: {json.dumps({'type': 'done'}, ensure_ascii=False)}\n\n"
            
        except Exception as e:
            logger.error(f"流式聊天处理失败: {str(e)}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'error': str(e)}, ensure_ascii=False)}\n\n"
    
    return StreamingResponse(generate(), media_type="text/event-stream")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
